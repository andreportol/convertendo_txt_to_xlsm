import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils.exceptions import IllegalCharacterError
import sys


def sanitizar_registros(registros: str) -> str:
    return ILLEGAL_CHARACTERS_RE.sub(r'?', registros)


def preencher_registros_na_planilha(sheet, registros: str):
    # PREENCHENDO OS REGISTROS na planilha
    cod_face = registros[0:8]
    A1 = sheet.cell(row=i, column=1)
    A1.value = cod_face
    cod_lote = registros[8:11]
    A2 = sheet.cell(row=i, column=2)
    A2.value = cod_lote
    cod_imo = registros[11:14]
    A3 = sheet.cell(row=i, column=3)
    A3.value = cod_imo
    digito = registros[14:15]
    A4 = sheet.cell(row=i, column=4)
    A4.value = digito
    insc_Anterior = registros[15:26]
    A5 = sheet.cell(row=i, column=5)
    A5.value = insc_Anterior
    face_lote = registros[26:34]
    A6 = sheet.cell(row=i, column=6)
    A6.value = face_lote
    cidade = registros[34:46]
    A7 = sheet.cell(row=i, column=7)
    A7.value = cidade
    tipo_log = registros[64:74]
    A8 = sheet.cell(row=i, column=8)
    A8.value = tipo_log
    end_imovel = registros[74:123]
    A9 = sheet.cell(row=i, column=9)
    A9.value = end_imovel
    num_porta = registros[124:130]
    A10 = sheet.cell(row=i, column=10)
    A10.value = num_porta
    complemento = registros[130:146]
    A11 = sheet.cell(row=i, column=11)
    A11.value = complemento
    parcelamento = registros[146:176]
    A12 = sheet.cell(row=i, column=12)
    A12.value = parcelamento
    quadra = registros[176:181]
    A13 = sheet.cell(row=i, column=13)
    A13.value = quadra
    lote = registros[181:186]
    A14 = sheet.cell(row=i, column=14)
    A14.value = lote
    rua_correspondencia = registros[186:216]
    A15 = sheet.cell(row=i, column=15)
    A15.value = rua_correspondencia
    nrPortaCorresp = registros[216:222]
    A16 = sheet.cell(row=i, column=16)
    A16.value = nrPortaCorresp
    complem_corresp = registros[222:238]
    A17 = sheet.cell(row=i, column=17)
    A17.value = complem_corresp
    bairroCorresp = registros[238:268]
    A18 = sheet.cell(row=i, column=18)
    A18.value = bairroCorresp
    cidadeCorresp = registros[268:298]
    A19 = sheet.cell(row=i, column=19)
    A19.value = cidadeCorresp
    estadoCorresp = registros[298:300]
    A20 = sheet.cell(row=i, column=20)
    A20.value = estadoCorresp
    cepCorresp = registros[300:309]
    A21 = sheet.cell(row=i, column=21)
    A21.value = cepCorresp
    proprietarioCorresp = registros[309:349]
    A22 = sheet.cell(row=i, column=22)
    A22.value = proprietarioCorresp
    compromissario = registros[349:389]
    A23 = sheet.cell(row=i, column=23)
    A23.value = compromissario
    administrador = registros[389:429]
    A24 = sheet.cell(row=i, column=24)
    A24.value = administrador
    patrimonio = registros[429:444]
    A25 = sheet.cell(row=i, column=25)
    A25.value = patrimonio
    taxacao = registros[444:459]
    A26 = sheet.cell(row=i, column=26)
    A26.value = taxacao
    descricao = registros[459:474]
    A27 = sheet.cell(row=i, column=27)
    A27.value = descricao
    uso_imovel = registros[474:489]
    A28 = sheet.cell(row=i, column=28)
    A28.value = uso_imovel
    cod_averbacao = registros[489:499]
    A29 = sheet.cell(row=i, column=29)
    A29.value = cod_averbacao
    agua = registros[499:500]
    A30 = sheet.cell(row=i, column=30)
    A30.value = agua
    coleta_lixo = registros[500:501]
    A31 = sheet.cell(row=i, column=31)
    A31.value = coleta_lixo
    coletivo = registros[501:502]
    A32 = sheet.cell(row=i, column=32)
    A32.value = coletivo
    esgoto = registros[502:503]
    A33 = sheet.cell(row=i, column=33)
    A33.value = esgoto
    ilum_public = registros[503:504]
    A34 = sheet.cell(row=i, column=34)
    A34.value = ilum_public
    limpeza = registros[504:505]
    A35 = sheet.cell(row=i, column=35)
    A35.value = limpeza
    meio_fio = registros[505:506]
    A36 = sheet.cell(row=i, column=36)
    A36.value = meio_fio
    pavimentacao = registros[506:507]
    A37 = sheet.cell(row=i, column=37)
    A37.value = pavimentacao
    rede_eletrica = registros[507:508]
    A38 = sheet.cell(row=i, column=38)
    A38.value = rede_eletrica
    rede_telefonica = registros[508:509]
    A39 = sheet.cell(row=i, column=39)
    A39.value = rede_telefonica
    setor_calculo = registros[509:515]
    A40 = sheet.cell(row=i, column=40)
    A40.value = setor_calculo


wb = openpyxl.Workbook()
sheet = wb.active
# INSERINDO NOME NAS COLUNAS
A1 = sheet.cell(row=1, column=1)
A1.value="COD.FACE"
A2 = sheet.cell(row=1, column=2)
A2.value="COD.LOTE"
A3 = sheet.cell(row=1, column=3)
A3.value="COD.IMO"
A4 = sheet.cell(row=1, column=4)
A4.value="DIGITO"
A5 = sheet.cell(row=1, column=5)
A5.value="INSCANT"
A6 = sheet.cell(row=1, column=6)
A6.value="FACELOC"
A7 = sheet.cell(row=1, column=7)
A7.value="CIDADE"
A8 = sheet.cell(row=1, column=8)
A8.value="TIPOLOG"
A9 = sheet.cell(row=1, column=9)
A9.value="END. IMOVEL"
A10 = sheet.cell(row=1, column=10)
A10.value="NR.PORTA"
A11 = sheet.cell(row=1, column=11)
A11.value="COMPLEMENTO"
A12 = sheet.cell(row=1, column=12)
A12.value="PARCELAMENTO"
A13 = sheet.cell(row=1, column=13)
A13.value="QUADRA"
A14 = sheet.cell(row=1, column=14)
A14.value="LOTE"
A15 = sheet.cell(row=1, column=15)
A15.value="RUACORRESP"
A16 = sheet.cell(row=1, column=16)
A16.value="NRPORTACORRESP"
A17 = sheet.cell(row=1, column=17)
A17.value="COMPLCORRESP"
A18 = sheet.cell(row=1, column=18)
A18.value="BAIRROCORRESP"
A19 = sheet.cell(row=1, column=19)
A19.value="CIDADECORRESP"
A20 = sheet.cell(row=1, column=20)
A20.value="ESTADOCORRESP"
A21 = sheet.cell(row=1, column=21)
A21.value="CEPCORRESP"
A22 = sheet.cell(row=1, column=22)
A22.value="PROPRIETARIO"
A23 = sheet.cell(row=1, column=23)
A23.value="COMPROMISSARIO"
A24 = sheet.cell(row=1, column=24)
A24.value="ADMINISTRADOR"
A25 = sheet.cell(row=1, column=25)
A25.value="PATRIMONIO"
A26 = sheet.cell(row=1, column=26)
A26.value="TAXACAO"
A27 = sheet.cell(row=1, column=27)
A27.value="DESCRICAO"
A28 = sheet.cell(row=1, column=28)
A28.value="USOIMOVEL"
A29 = sheet.cell(row=1, column=29)
A29.value="CODAVERBACAO"
A30 = sheet.cell(row=1, column=30)
A30.value="AGUA"
A31 = sheet.cell(row=1, column=31)
A31.value="COLETALIXO"
A32 = sheet.cell(row=1, column=32)
A32.value="COLETIVO"
A33 = sheet.cell(row=1, column=33)
A33.value="ESGOTO"
A34 = sheet.cell(row=1, column=34)
A34.value="ILUMPUBLIC"
A35 = sheet.cell(row=1, column=35)
A35.value="LIMPEZA"
A36 = sheet.cell(row=1, column=36)
A36.value="MEIOFIO"
A37 = sheet.cell(row=1, column=37)
A37.value="PAVIMENTACAO"
A38 = sheet.cell(row=1, column=38)
A38.value="REDEELETRICA"
A39 = sheet.cell(row=1, column=39)
A39.value="REDETELEFONICA"
A40 = sheet.cell(row=1, column=40)
A40.value="SETORCALCULO"
A41 = sheet.cell(row=1, column=41)
A41.value="CODREGIAO"
A42 = sheet.cell(row=1, column=42)
A42.value="SETORREGIAO"
A43 = sheet.cell(row=1, column=43)
A43.value="FTKSET"
A44 = sheet.cell(row=1, column=44)
A44.value="TESTADA"
A45 = sheet.cell(row=1, column=45)
A45.value="NRTESTADA"
A46 = sheet.cell(row=1, column=46)
A46.value="AREA_TERRENO"
A47 = sheet.cell(row=1, column=47)
A47.value="AREA_LOTE"
A48 = sheet.cell(row=1, column=48)
A48.value="AREA_PISCINA"
A49 = sheet.cell(row=1, column=49)
A49.value="QUADRAESPORTE"
A50 = sheet.cell(row=1, column=50)
A50.value="TIPOLOTE"
A51 = sheet.cell(row=1, column=51)
A51.value="FORMATOTERRENO"
A52 = sheet.cell(row=1, column=52)
A52.value="CARACTLIMITE"
A53 = sheet.cell(row=1, column=53)
A53.value="TOPOGRAFIA"
A54 = sheet.cell(row=1, column=54)
A54.value="PEDOLOGIA"
A55 = sheet.cell(row=1, column=55)
A55.value="FRENTE"
A56 = sheet.cell(row=1, column=56)
A56.value="CALCADA"
A57 = sheet.cell(row=1, column=57)
A57.value="AREAEDIFCA"
A58 = sheet.cell(row=1, column=58)
A58.value="VALORVENALTERRENO"
A59 = sheet.cell(row=1, column=59)
A59.value="VALORCONSTRUCA"
A60 = sheet.cell(row=1, column=60)
A60.value="AVALIACAO"
A61 = sheet.cell(row=1, column=61)
A61.value="ALIQUOTA"
A62 = sheet.cell(row=1, column=62)
A62.value="FRACAOIDEAL"
A63 = sheet.cell(row=1, column=63)
A63.value="DATAALTERACAO"        
    
# ABRINDO ARQUIVO TXT
arquivo = open('teste.txt')
dados = arquivo.read()
arquivo.close()
# LENDO ARQUIVO
i = 2


for registros in dados.splitlines():
    try:
        preencher_registros_na_planilha(sheet, registros)
    except IllegalCharacterError:
        registros_sanitizado = sanitizar_registros(registros)
        '''
        print(f"Caracteres ilegais foram encontrados e substituidos por '?' na linha:\n{registros}\n"
              f"Esta linha foi substituida por:\n{registros_sanitizado}\n"
              f"Verifique esta substituicao\n", file=sys.stderr)
        '''
        preencher_registros_na_planilha(sheet, registros_sanitizado)
    i += 1
   
     
   
wb.save(r"C:\Users\andre.porto\Desktop\arquivostique\controle.xlsx")